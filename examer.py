"""
# 30.10.2023
Developer: Mehdi Safarzade (SeonerVorteX - https://github.com/SeonerVorteX/)
Requirements: pip install -r requirements.txt
How to use: python examer.py and follow the instructions
Note: 
    You can add your own exams to the ./templates folder. The exam template should be in the form of a .xlsx file. 
    The first column should be the question, the second column should be the first variant, the third column should 
    be the second variant, the fourth column should be the third variant, the fifth column should be the fourth variant, 
    the sixth column should be the fifth variant, and the seventh column should be the correct answer. You can use the 
    Example.xlsx file in the ./templates folder as an example.
"""

import os
import glob
import time
import random
from openpyxl import Workbook,load_workbook

header = "----------------EXAMER----------------\n"
commands = ['start_exam', 'exit']
examFiles = glob.glob('templates/*.xlsx')
exams = [os.path.basename(file).replace(".xlsx", "").upper() for file in examFiles]

print(header)

if  __name__ == '__main__':
    print('Commands: {}\n'.format(' | '.join(commands)))
    
    while True:
        command = input('/> ').lower().strip()
        if not(command):
            continue
        elif command == 'exit':
            break
        elif command == "start_exam":
            success = False
            test = ""
            print('Exams: {}\n'.format(' | '.join(exams)))
            while not(success):
                test = input('Choose you exam: ').upper().strip()
                if not(test) or not(test in exams):
                    print('Please enter a valid exam\n')
                else:
                    success = True
                    break
    
            wb = load_workbook('templates/{}.xlsx'.format(test))
            ws = wb.active
            print("Exam starts...\n")
            time.sleep(1.5)
            success = False
            count = 0
            while not(success):
                try:
                    question_count = input("How many questions do you want to answer? (Enter 'exit' to exit): ")
                    if question_count.lower() == 'exit':
                        success = True
                        break
                    else:
                        count = int(question_count)
                except ValueError:
                    print('Please enter a number\n')
                    continue
                if count < 0:
                    print('Please enter a positive number\n')
                    continue
                elif not(count):
                    print('Please enter a number\n')
                    continue
                elif count > ws.max_row-1:
                    print('Please enter a number less than or equal to {}\n'.format(ws.max_row-1))
                    continue
            
                questions = random.sample(range(2, ws.max_row+1),count)
                print('Exam started\n')
                time.sleep(1.5)
                start_time = time.time()
                correct_answers = 0
                incorrect_answers = 0
                question_row = 1
                for row in questions:
                    question = ws.cell(row=row,column=1).value
                    variant_a = ws.cell(row=row,column=2).value
                    variant_b = ws.cell(row=row,column=3).value
                    variant_c = ws.cell(row=row,column=4).value
                    variant_d = ws.cell(row=row,column=5).value
                    variant_e = ws.cell(row=row,column=6).value
            
                    variants = {
                        "A": variant_a,
                        "B": variant_b,
                        "C": variant_c,
                        "D": variant_d,
                        "E": variant_e
                    }
                    answer_variant = ws.cell(row=row,column=7).value.upper()
                    answer_content = variants[answer_variant]
                    
                    print('{}. {}?'.format(question_row,question))
                    print('A) {}\nB) {}\nC) {}\nD) {}\nE) {}\n'.format(variant_a,variant_b,variant_c,variant_d,variant_e))
                    
                    validation = False
                    while not(validation):
                        user_answer = input('Your answer: ').upper().strip()
                        if user_answer.lower() == 'exit':
                            validation = True
                            break
                        elif not(user_answer in variants):
                            print('Please enter a valid answer\n')
                            continue
                        else:
                            validation = True
                        
                    if user_answer.lower() == 'exit':
                        print('İmtahan bitdi\n')
                        success = True
                        break
                    elif user_answer == answer_variant:
                        print('Correct answer\n')
                        correct_answers += 1
                    else:
                        print('Wrong answer, Correct answer: {}\n'.format(answer_variant))
                        incorrect_answers += 1
                    time.sleep(1)
                    question_row += 1
                
                success = True
                end_time = time.time()
                time_taken = end_time-start_time
                minute, second = divmod(time_taken, 60)
                print('Exam finished\n')
                print('Correct answers: {}'.format(correct_answers))
                print('Wrong answers: {}'.format(incorrect_answers))
                print('Time taken: {} minutes {} seconds'.format(round(minute),round(second)))
        else:
            print('Enter a valid command\Commands: {}\n'.format(' | '.join(commands)))
            
